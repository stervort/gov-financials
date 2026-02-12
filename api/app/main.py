# api/app/main.py
from __future__ import annotations

import os
import io
import csv
from decimal import Decimal
from typing import Dict, Any, Optional

from fastapi import FastAPI, Request, UploadFile, Form, Query
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

from sqlalchemy import select, desc, func
from sqlalchemy.orm import Session

from openpyxl import load_workbook

from .db import engine, SessionLocal
from .models import (
    Base,
    UploadedFile,
    Fund,
    Account,
    TBImport,
    TBLine,
    Client,
    Binder,
)
from .tb_import import (
    read_csv_preview,
    ColumnMapping,
    validate_tb,
    iter_csv_rows,
    derive_fund_from_account,
    normalize_amount,
)

# -----------------------
# App + Templates
# -----------------------
app = FastAPI()
templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "templates"))

Base.metadata.create_all(bind=engine)


def db_session() -> Session:
    return SessionLocal()


# -----------------------
# Helpers (avoid detached ORM in Jinja)
# -----------------------
def load_binder_header(db: Session, binder_id: int) -> Dict[str, Any]:
    b = db.scalar(select(Binder).where(Binder.id == binder_id))
    if not b:
        return {"binder_id": None, "binder_label": None, "entity_name": None, "period_end": None}

    c = db.scalar(select(Client).where(Client.id == b.client_id)) if getattr(b, "client_id", None) else None

    period_end = getattr(b, "period_end", None)
    period_end_str = period_end.strftime("%m/%d/%Y") if period_end else ""

    binder_label = getattr(b, "binder_name", None)
    if not binder_label:
        entity_name = getattr(c, "client_name", "") if c else ""
        binder_label = f"{entity_name} ({period_end_str})".strip()

    return {
        "binder_id": b.id,
        "binder_label": binder_label,
        "entity_name": getattr(c, "client_name", None) if c else None,
        "period_end": period_end_str or None,
    }


def template_ctx(request: Request, **kwargs) -> Dict[str, Any]:
    ctx = {"request": request}
    ctx.update(kwargs)
    return ctx


def xlsx_bytes_to_csv_text(xlsx_bytes: bytes) -> str:
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    max_cols = max(len(r) for r in rows)

    def norm(v) -> str:
        if v is None:
            return ""
        return str(v)

    table = []
    for r in rows:
        row = [norm(r[i]) if i < len(r) else "" for i in range(max_cols)]
        table.append(row)

    # trim trailing fully-empty columns
    while max_cols > 1 and all((row[max_cols - 1] == "" for row in table)):
        max_cols -= 1
        table = [row[:max_cols] for row in table]

    out = io.StringIO()
    w = csv.writer(out)
    for row in table:
        w.writerow(row)
    return out.getvalue()


# -----------------------
# Health
# -----------------------
@app.get("/health")
def health():
    return {"status": "ok"}


# ============================================================
# HOME (Select Binder)
# ============================================================
@app.get("/", response_class=HTMLResponse)
def home(request: Request, client_id: Optional[int] = Query(default=None)):
    with db_session() as db:
        clients = db.execute(select(Client).order_by(Client.client_name)).scalars().all()
        client_rows = [{"id": c.id, "client_name": c.client_name} for c in clients]

        selected_client = None
        binder_rows = []

        if client_id:
            selected_client = db.scalar(select(Client).where(Client.id == client_id))

            binders = db.execute(
                select(Binder).where(Binder.client_id == client_id).order_by(desc(Binder.period_end), desc(Binder.id))
            ).scalars().all()

            for b in binders:
                period_end = getattr(b, "period_end", None)
                period_end_str = period_end.strftime("%m/%d/%Y") if period_end else ""

                binder_label = getattr(b, "binder_name", None)
                if not binder_label:
                    binder_label = f"{selected_client.client_name} ({period_end_str})".strip() if selected_client else ""

                binder_rows.append(
                    {
                        "binder_id": b.id,
                        "binder_label": binder_label,
                        "period_end": period_end_str,
                    }
                )

    return templates.TemplateResponse(
        "home.html",
        template_ctx(
            request,
            clients=client_rows,
            selected_client_id=client_id,
            selected_entity_name=(selected_client.client_name if selected_client else None),
            binders=binder_rows,
            message=None,
        ),
    )


# ============================================================
# NEW ENTITY WIZARD (creates entity + first binder)
# ============================================================
@app.get("/entity/new", response_class=HTMLResponse)
def new_entity_page(request: Request):
    return templates.TemplateResponse("new_entity.html", template_ctx(request, message=None))


@app.post("/entity/new")
def new_entity_create(
    request: Request,
    entity_name: str = Form(...),
    address: str = Form(""),
    contact_name: str = Form(""),
    contact_email: str = Form(""),
    period_end: str = Form(...),  # YYYY-MM-DD
):
    from datetime import date

    y, m, d = [int(x) for x in period_end.split("-")]
    pe = date(y, m, d)

    with db_session() as db:
        c = Client(
            client_name=entity_name.strip(),
            address=(address or "").strip(),
            contact_name=(contact_name or "").strip(),
            contact_email=(contact_email or "").strip(),
        )
        db.add(c)
        db.flush()

        binder_label = f"{c.client_name} ({pe.strftime('%m/%d/%Y')})"

        b = Binder(
            client_id=c.id,
            period_end=pe,
            binder_name=binder_label,
        )
        db.add(b)
        db.commit()
        db.refresh(b)

    return RedirectResponse(f"/binder/{b.id}/open", status_code=303)


# ============================================================
# BINDER ACTIONS
# ============================================================
@app.get("/binder/{binder_id}/open")
def binder_open(request: Request, binder_id: int):
    from datetime import datetime

    with db_session() as db:
        b = db.get(Binder, binder_id)
        if not b:
            return RedirectResponse("/", status_code=303)

        if hasattr(b, "last_accessed"):
            b.last_accessed = datetime.utcnow()
            db.commit()

    return RedirectResponse(f"/binder/{binder_id}/dashboard", status_code=303)


@app.get("/binder/{binder_id}/dashboard", response_class=HTMLResponse)
def binder_dashboard(request: Request, binder_id: int):
    with db_session() as db:
        header = load_binder_header(db, binder_id)
        if not header["binder_id"]:
            return RedirectResponse("/", status_code=303)

        latest = db.scalar(
            select(TBImport).where(TBImport.binder_id == binder_id).order_by(desc(TBImport.created_at))
        )

        latest_import_name = latest.import_name if latest else None
        latest_import_date = latest.created_at.strftime("%m/%d/%Y %H:%M") if latest and latest.created_at else None

        latest_import_lines = 0
        if latest:
            latest_import_lines = db.scalar(
                select(func.count()).select_from(TBLine).where(TBLine.tb_import_id == latest.id)
            ) or 0

    return templates.TemplateResponse(
        "binder_dashboard.html",
        template_ctx(
            request,
            **header,
            latest_import_name=latest_import_name,
            latest_import_date=latest_import_date,
            latest_import_lines=latest_import_lines,
        ),
    )


@app.get("/binder/{binder_id}/delete", response_class=HTMLResponse)
def binder_delete_confirm(request: Request, binder_id: int):
    with db_session() as db:
        header = load_binder_header(db, binder_id)
        if not header["binder_id"]:
            return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse(
        "binder_delete_confirm.html",
        template_ctx(request, **header, message=None),
    )


@app.post("/binder/{binder_id}/delete")
def binder_delete_do(
    request: Request,
    binder_id: int,
    confirm_text: str = Form(""),
):
    if confirm_text.strip() != "Yes I want to delete this binder":
        return RedirectResponse(f"/binder/{binder_id}/delete", status_code=303)

    with db_session() as db:
        imports = db.execute(select(TBImport).where(TBImport.binder_id == binder_id)).scalars().all()
        import_ids = [i.id for i in imports]

        if import_ids:
            db.query(TBLine).filter(TBLine.tb_import_id.in_(import_ids)).delete(synchronize_session=False)
            db.query(TBImport).filter(TBImport.id.in_(import_ids)).delete(synchronize_session=False)

        if hasattr(UploadedFile, "binder_id"):
            db.query(UploadedFile).filter(UploadedFile.binder_id == binder_id).delete(synchronize_session=False)

        b = db.get(Binder, binder_id)
        if b:
            db.delete(b)

        db.commit()

    return RedirectResponse("/", status_code=303)


@app.get("/binder/{binder_id}/rollforward", response_class=HTMLResponse)
def binder_rollforward_page(request: Request, binder_id: int):
    with db_session() as db:
        header = load_binder_header(db, binder_id)
        if not header["binder_id"]:
            return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse("rollforward.html", template_ctx(request, **header, message=None))


@app.post("/binder/{binder_id}/rollforward")
def binder_rollforward_create(
    request: Request,
    binder_id: int,
    new_period_end: str = Form(...),  # YYYY-MM-DD
):
    from datetime import date

    y, m, d = [int(x) for x in new_period_end.split("-")]
    pe = date(y, m, d)

    with db_session() as db:
        old = db.get(Binder, binder_id)
        if not old:
            return RedirectResponse("/", status_code=303)

        c = db.get(Client, old.client_id) if getattr(old, "client_id", None) else None
        entity_name = c.client_name if c else "Entity"

        binder_label = f"{entity_name} ({pe.strftime('%m/%d/%Y')})"

        new_b = Binder(
            client_id=old.client_id,
            period_end=pe,
            binder_name=binder_label,
        )
        db.add(new_b)
        db.commit()
        db.refresh(new_b)

    return RedirectResponse(f"/binder/{new_b.id}/open", status_code=303)


# ============================================================
# TB Import Wizard (XLSX ONLY) – binder-scoped
# ============================================================
@app.get("/tb/upload", response_class=HTMLResponse)
def tb_upload_page(request: Request, binder_id: int = Query(...)):
    with db_session() as db:
        header = load_binder_header(db, binder_id)
        if not header["binder_id"]:
            return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse("upload.html", template_ctx(request, **header, message=None))


@app.post("/tb/upload", response_class=HTMLResponse)
async def tb_upload(
    request: Request,
    binder_id: int = Form(...),
    file: UploadFile = Form(...),
):
    raw = await file.read()
    filename = file.filename or "upload.xlsx"
    lower = filename.lower()

    with db_session() as db:
        header = load_binder_header(db, binder_id)
        if not header["binder_id"]:
            return RedirectResponse("/", status_code=303)

    if not lower.endswith(".xlsx"):
        return templates.TemplateResponse(
            "upload.html",
            template_ctx(request, **header, message="Please upload an Excel .xlsx trial balance file."),
        )

    content = xlsx_bytes_to_csv_text(raw)

    has_headers_bool = True
    delimiter_used = ","
    header_row_used = 1
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    with db_session() as db:
        uf = UploadedFile(
            filename=filename,
            content_type=content_type,
            content_text=content,
        )
        if hasattr(uf, "binder_id"):
            uf.binder_id = binder_id

        db.add(uf)
        db.commit()
        db.refresh(uf)

    headers, rows = read_csv_preview(
        content,
        has_headers=has_headers_bool,
        header_row=header_row_used,
        delimiter=delimiter_used,
        max_rows=50,
    )

    return templates.TemplateResponse(
        "map.html",
        template_ctx(
            request,
            **header,
            binder_id=binder_id,
            upload_id=uf.id,
            filename=filename,
            headers=headers,
            rows=rows,
            has_headers=True,
            delimiter=",",
            header_row=1,
            message=None,
        ),
    )


@app.post("/tb/validate", response_class=HTMLResponse)
def tb_validate(
    request: Request,
    binder_id: int = Form(...),
    upload_id: int = Form(...),
    has_headers: int = Form(...),
    delimiter: str = Form(...),
    header_row: int = Form(...),

    account_col: str = Form(...),
    desc_col: str = Form(...),

    amount_mode: str = Form(...),
    balance_col: str = Form(""),
    debit_col: str = Form(""),
    credit_col: str = Form(""),
    credit_sign_mode: str = Form("keep"),

    fund_mode: str = Form(...),
    fund_col: str = Form(""),
    fund_delimiter: str = Form("-"),
):
    with db_session() as db:
        header = load_binder_header(db, binder_id)
        uf = db.get(UploadedFile, upload_id)
        csv_text = uf.content_text if uf else ""

    mapping = ColumnMapping(
        account_col=account_col,
        desc_col=desc_col.strip() if desc_col.strip() else None,
        mode=amount_mode,
        balance_col=balance_col if amount_mode == "signed" else None,
        debit_col=debit_col if amount_mode == "dc" else None,
        credit_col=credit_col if amount_mode == "dc" else None,
        credit_sign_mode=(credit_sign_mode or "keep").strip().lower(),
        fund_mode=fund_mode,
        fund_col=fund_col if fund_mode == "fund_column" else None,
        fund_delimiter=fund_delimiter or "-",
    )

    report = validate_tb(
        csv_text=csv_text,
        has_headers=bool(has_headers),
        header_row=int(header_row),
        delimiter=delimiter,
        mapping=mapping,
    )

    tolerance = Decimal("1.00")
    nets_to_zero = abs(Decimal(str(report["total_net"]))) <= tolerance

    return templates.TemplateResponse(
        "validate.html",
        template_ctx(
            request,
            **header,
            binder_id=binder_id,
            upload_id=upload_id,
            mapping=mapping,
            has_headers=bool(has_headers),
            delimiter=delimiter,
            header_row=int(header_row),
            report=report,
            tolerance=tolerance,
            nets_to_zero=nets_to_zero,
            force_unbalanced_warning=False,
        ),
    )


@app.post("/tb/funds", response_class=HTMLResponse)
def tb_funds(
    request: Request,
    binder_id: int = Form(...),
    upload_id: int = Form(...),

    has_headers: int = Form(...),
    delimiter: str = Form(...),
    header_row: int = Form(...),

    account_col: str = Form(...),
    desc_col: str = Form(""),
    amount_mode: str = Form(...),
    balance_col: str = Form(""),
    debit_col: str = Form(""),
    credit_col: str = Form(""),
    credit_sign_mode: str = Form("keep"),

    fund_mode: str = Form(...),
    fund_col: str = Form(""),
    fund_delimiter: str = Form("-"),

    allow_unbalanced: str = Form("off"),
):
    with db_session() as db:
        header = load_binder_header(db, binder_id)
        uf = db.get(UploadedFile, upload_id)
        csv_text = uf.content_text if uf else ""

    mapping = ColumnMapping(
        account_col=account_col,
        desc_col=desc_col.strip() if desc_col.strip() else None,
        mode=amount_mode,
        balance_col=balance_col if amount_mode == "signed" else None,
        debit_col=debit_col if amount_mode == "dc" else None,
        credit_col=credit_col if amount_mode == "dc" else None,
        credit_sign_mode=(credit_sign_mode or "keep").strip().lower(),
        fund_mode=fund_mode,
        fund_col=fund_col if fund_mode == "fund_column" else None,
        fund_delimiter=fund_delimiter or "-",
    )

    report = validate_tb(
        csv_text=csv_text,
        has_headers=bool(has_headers),
        header_row=int(header_row),
        delimiter=delimiter,
        mapping=mapping,
    )

    tolerance = Decimal("1.00")
    nets_to_zero = abs(Decimal(str(report["total_net"]))) <= tolerance

    if (not nets_to_zero) and allow_unbalanced != "on":
        return templates.TemplateResponse(
            "validate.html",
            template_ctx(
                request,
                **header,
                binder_id=binder_id,
                upload_id=upload_id,
                mapping=mapping,
                has_headers=bool(has_headers),
                delimiter=delimiter,
                header_row=int(header_row),
                report=report,
                tolerance=tolerance,
                nets_to_zero=nets_to_zero,
                force_unbalanced_warning=True,
            ),
        )

    fund_codes = sorted(report["fund_counts"].keys())

    return templates.TemplateResponse(
        "funds.html",
        template_ctx(
            request,
            **header,
            binder_id=binder_id,
            upload_id=upload_id,
            has_headers=int(has_headers),
            delimiter=delimiter,
            header_row=int(header_row),
            mapping=mapping,
            fund_codes=fund_codes,
            fund_counts=report["fund_counts"],
        ),
    )


@app.post("/tb/import", response_class=HTMLResponse)
async def tb_import_commit(request: Request):
    form = dict(await request.form())

    binder_id = int(form["binder_id"])
    upload_id = int(form["upload_id"])
    has_headers = bool(int(form["has_headers"]))
    delimiter = form["delimiter"]
    header_row = int(form["header_row"])

    account_col = form["account_col"]
    desc_col = (form.get("desc_col") or "").strip()

    amount_mode = form["amount_mode"]
    balance_col = (form.get("balance_col") or "").strip()
    debit_col = (form.get("debit_col") or "").strip()
    credit_col = (form.get("credit_col") or "").strip()
    credit_sign_mode = (form.get("credit_sign_mode") or "keep").strip().lower()

    fund_mode = form["fund_mode"]
    fund_col = (form.get("fund_col") or "").strip()
    fund_delimiter = (form.get("fund_delimiter") or "-").strip()

    mapping = ColumnMapping(
        account_col=account_col,
        desc_col=desc_col if desc_col else None,
        mode=amount_mode,
        balance_col=balance_col if amount_mode == "signed" else None,
        debit_col=debit_col if amount_mode == "dc" else None,
        credit_col=credit_col if amount_mode == "dc" else None,
        credit_sign_mode=credit_sign_mode,
        fund_mode=fund_mode,
        fund_col=fund_col if fund_mode == "fund_column" else None,
        fund_delimiter=fund_delimiter or "-",
    )

    imported_lines = 0
    uf_filename = "upload.xlsx"

    with db_session() as db:
        header = load_binder_header(db, binder_id)
        if not header["binder_id"]:
            return RedirectResponse("/", status_code=303)

        uf = db.get(UploadedFile, upload_id)
        if not uf:
            return templates.TemplateResponse(
                "complete.html",
                template_ctx(
                    request,
                    **header,
                    import_name="Import failed",
                    imported_lines=0,
                ),
            )

        uf_filename = uf.filename
        csv_text = uf.content_text

        tbi = TBImport(
            import_name=f"TB Import - {uf_filename}",
            uploaded_file_id=upload_id,
            binder_id=binder_id,
        )
        db.add(tbi)
        db.commit()
        db.refresh(tbi)

        fund_cache: Dict[str, Fund] = {}
        for k, v in form.items():
            if k.startswith("fund_name__"):
                code = k.split("__", 1)[1]
                name = str(v).strip()
                ftype = str(form.get(f"fund_type__{code}", "")).strip()

                existing = db.scalar(select(Fund).where(Fund.fund_code == code))
                if not existing:
                    existing = Fund(fund_code=code, fund_name=name, fund_type=ftype)
                    db.add(existing)
                    db.flush()
                else:
                    existing.fund_name = name
                    existing.fund_type = ftype

                fund_cache[code] = existing

        db.flush()

        acct_cache: Dict[str, Account] = {}

        for rowno, row in iter_csv_rows(csv_text, has_headers, header_row, delimiter):
            acct = (row.get(mapping.account_col, "") or "").strip()
            if mapping.ignore_blank_account and acct == "":
                continue

            if mapping.fund_mode == "fund_from_account_prefix":
                fund_code = derive_fund_from_account(acct, mapping.fund_delimiter)
            elif mapping.fund_mode == "fund_column":
                fund_code = (row.get(mapping.fund_col or "", "") or "").strip()
            else:
                fund_code = "SINGLE"

            if fund_code == "":
                continue

            amt, _warns = normalize_amount(row, mapping)
            if amt is None:
                continue
            if mapping.ignore_zero and amt == 0:
                continue

            desc = ""
            if mapping.desc_col:
                desc = (row.get(mapping.desc_col, "") or "").strip()

            fund = fund_cache.get(fund_code) or db.scalar(select(Fund).where(Fund.fund_code == fund_code))
            if not fund:
                fund = Fund(fund_code=fund_code, fund_name="", fund_type="")
                db.add(fund)
                db.flush()
                fund_cache[fund_code] = fund

            acct_obj = acct_cache.get(acct) or db.scalar(select(Account).where(Account.account_number == acct))
            if not acct_obj:
                acct_obj = Account(account_number=acct, account_name=desc[:255] if desc else "")
                db.add(acct_obj)
                db.flush()
                acct_cache[acct] = acct_obj

            line = TBLine(
                tb_import_id=tbi.id,
                binder_id=binder_id,
                fund_id=fund.id,
                account_id=acct_obj.id,
                description=desc[:255],
                amount=float(amt),
                source_row=rowno,
            )
            db.add(line)
            imported_lines += 1

        db.commit()

    return templates.TemplateResponse(
        "complete.html",
        template_ctx(
            request,
            **header,
            binder_id=binder_id,
            import_name=f"TB Import - {uf_filename}",
            imported_lines=imported_lines,
        ),
    )
